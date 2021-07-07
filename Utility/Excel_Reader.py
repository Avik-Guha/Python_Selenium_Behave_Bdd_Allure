import openpyxl


def read_excel_data(file_path, cell_name, column):
    # logger = Base.set_Logger()

    # Load Workbook
    workbook_ob = openpyxl.load_workbook(file_path)

    # Create object of the sheet you want to work on
    sheet1_ob = workbook_ob['Sheet1']

    # Find Rows having data
    rows = sheet1_ob.max_row
    columns = sheet1_ob.max_column

    profile_index = 0
    username_index = 0
    password_index = 0
    email_index = 0
    user_row = 0

    profile = cell_name

    # Find index values
    for i in range(1, 2):
        for j in range(1, columns + 1):
            cell = sheet1_ob.cell(i, j)
            if cell.value == "Profile":
                profile_index = j
            elif cell.value == "Username":
                username_index = j
            elif cell.value == "Password":
                password_index = j
            elif cell.value == "Email":
                email_index = j

    # Locate row for required user
    for i in range(1, rows + 1):
        cell = sheet1_ob.cell(i, profile_index)
        if cell.value == profile:
            user_row = i
            break

    if column == "Username":
        # logger.info("Fetch Username")
        username = sheet1_ob.cell(user_row, username_index)
        req_value = username.value
    elif column == "Password":
        # logger.info("Fetch Password")
        password = sheet1_ob.cell(user_row, password_index)
        req_value = password.value
    elif column == "Email":
        # logger.info("Fetch Email")
        email = sheet1_ob.cell(user_row, email_index)
        req_value = email.value

    # logger.info("Returning value from excel: " + req_value)
    return req_value

